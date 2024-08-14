from fastapi import HTTPException, Request
from shutil import copyfileobj
import os
import openpyxl as xl
from fpdf import FPDF


def path_and_sheets(file):
    os.makedirs('files', exist_ok=True)
    path = f"files/{file.filename}"
    try:
        with open(path, "wb+") as file_object:
            copyfileobj(file.file, file_object)
        exc = xl.load_workbook(path)
    except:
        raise HTTPException(status_code=400, detail="An error occurred while processing the file")

    # including funcs of part2
    return {"filename": file.filename, "fileRouting": os.path.join(os.getcwd(), ),
            "sheets": len(exc.sheetnames)}


def process_sheet(sheet_data, worksheet):
    operation = sheet_data.get('operation')
    columns = sheet_data.get('columns')
    if operation == "sum":
        results = []
        for column in columns:
            values = [cell.value for cell in worksheet[column] if cell.value is not None]
            total_sum = sum(float(value) for value in values if value is not None)
            results.append({"sum of column:" + column: total_sum})
        return results

    elif operation == "avg":
        results = []
        for column in columns:
            values = [cell.value for cell in worksheet[column] if cell.value is not None]
            values = [float(value) for value in values if value is not None]
            avg_value = sum(values) / len(values) if len(values) > 0 else 0
            results.append({"avg of column:" + column: avg_value})
        return results

    else:
        return []


def export_json(request_data):
    path = request_data.get('path')
    report_data = {}

    try:
        workbook = xl.load_workbook(path)
        for sheet_data in request_data.get('sheets', []):
            worksheet = workbook[sheet_data.get('name')]
            report_data[sheet_data.get('name')] = process_sheet(sheet_data, worksheet)
        workbook.close()
        return report_data

    except Exception:
        error_message = "An error occurred while generating the report."
        raise HTTPException(status_code=500, detail=error_message)


def generate_pdf_from_json(data):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.add_page()
    pdf.set_font("Arial", size=12)
    for key, value in data.items():
        pdf.cell(200, 10, f"{key}: {value}", ln=True)

    return pdf


async def pdf_from_json(request: Request):
    data = await request.json()

    pdf = generate_pdf_from_json(data)
    pdf_output = pdf.output(dest='S').encode('latin1')

    # Save the PDF file
    filename = "output.pdf"
    with open(filename, 'wb') as output_file:
        output_file.write(pdf_output)

    return {"message": "PDF file created successfully", "filename": filename}
